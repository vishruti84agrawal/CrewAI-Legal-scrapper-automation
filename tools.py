"""
Web scraping tools for ATTOM Data Collection
Provides browser automation and data extraction capabilities
"""

from crewai.tools import BaseTool
from typing import Type, Dict, List, Any, Optional
from pydantic import BaseModel, Field
import requests
from bs4 import BeautifulSoup
import time
import random
import json
import time
import json
import re
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime
import os


class WebScraperInput(BaseModel):
    """Input schema for web scraping operations."""
    url: str = Field(..., description="The URL to scrape")
    wait_time: int = Field(default=3, description="Time to wait after loading page (seconds)")
    css_selectors: Optional[Dict[str, str]] = Field(default=None, description="CSS selectors for data extraction")


class DataExtractorInput(BaseModel):
    """Input schema for data extraction."""
    html_content: str = Field(..., description="HTML content to extract data from")
    extraction_rules: Dict[str, str] = Field(..., description="Field name to CSS selector mapping")


class BasicWebScrapingTool(BaseTool):
    """Basic web scraping tool using requests and BeautifulSoup."""
    
    name: str = "web_scraper"
    description: str = "Scrape web pages and extract data using CSS selectors"
    args_schema: Type[BaseModel] = WebScraperInput

    def _run(self, url: str, wait_time: int = 3, css_selectors: Dict[str, str] = None) -> str:
        try:
            # Enhanced headers to avoid detection
            headers = {
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
                'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
                'Accept-Language': 'en-US,en;q=0.5',
                'Accept-Encoding': 'gzip, deflate, br',
                'Connection': 'keep-alive',
                'Upgrade-Insecure-Requests': '1',
                'Sec-Fetch-Dest': 'document',
                'Sec-Fetch-Mode': 'navigate',
                'Sec-Fetch-Site': 'none',
                'Cache-Control': 'max-age=0'
            }
            
            # Add random delay to avoid rate limiting
            time.sleep(wait_time + random.uniform(1, 3))
            
            # Create session for cookie persistence
            session = requests.Session()
            session.headers.update(headers)
            
            response = requests.get(url, headers=headers)
            response.raise_for_status()
            
            soup = BeautifulSoup(response.content, 'html.parser')
            
            if css_selectors:
                # Extract specific data based on selectors
                extracted_data = {}
                for field_name, selector in css_selectors.items():
                    elements = soup.select(selector)
                    if elements:
                        extracted_data[field_name] = [elem.get_text(strip=True) for elem in elements]
                    else:
                        extracted_data[field_name] = []
                
                return json.dumps({
                    "status": "success",
                    "url": url,
                    "extracted_data": extracted_data
                }, indent=2)
            else:
                # Return page title and basic info
                title = soup.find('title').get_text(strip=True) if soup.find('title') else "No title"
                
                return json.dumps({
                    "status": "success",
                    "url": url,
                    "title": title,
                    "page_text_preview": soup.get_text()[:500] + "..." if len(soup.get_text()) > 500 else soup.get_text()
                }, indent=2)
                
        except Exception as e:
            return json.dumps({
                "status": "error",
                "url": url,
                "error": str(e)
            }, indent=2)


class NYSCourtNavigationTool(BaseTool):
    """Specialized tool for navigating NYS Unified Court System."""
    
    name: str = "nys_court_navigator"
    description: str = "Navigate NYS Unified Court System for foreclosure cases"
    args_schema: Type[BaseModel] = WebScraperInput

    def _run(self, url: str = "https://iapps.courts.state.ny.us/nyscef/CaseSearch", 
             wait_time: int = 3, css_selectors: Dict[str, str] = None) -> str:
        try:
            # Schenectady County specific navigation
            headers = {
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
            }
            
            # Step 1: Access the court system
            response = requests.get(url, headers=headers)
            response.raise_for_status()
            
            soup = BeautifulSoup(response.content, 'html.parser')
            
            # Look for foreclosure case links and forms
            foreclosure_links = []
            search_forms = []
            
            # Find foreclosure-related links
            for link in soup.find_all('a', href=True):
                link_text = link.get_text(strip=True).lower()
                if any(keyword in link_text for keyword in ['foreclosure', 'notice of pendency', 'lis pendens']):
                    foreclosure_links.append({
                        "text": link.get_text(strip=True),
                        "href": link['href']
                    })
            
            # Find search forms
            for form in soup.find_all('form'):
                form_inputs = []
                for inp in form.find_all(['input', 'select']):
                    form_inputs.append({
                        "type": inp.get('type', 'text'),
                        "name": inp.get('name', ''),
                        "id": inp.get('id', ''),
                        "placeholder": inp.get('placeholder', '')
                    })
                
                if form_inputs:
                    search_forms.append({
                        "action": form.get('action', ''),
                        "method": form.get('method', 'GET'),
                        "inputs": form_inputs
                    })
            
            return json.dumps({
                "status": "success",
                "site": "NYS Unified Court System",
                "county": "Schenectady",
                "foreclosure_links": foreclosure_links,
                "search_forms": search_forms,
                "instructions": "Next steps: Use 'SEARCH AS GUEST' → 'NEW CASES' → 'SCHENECTADY SUPREME COURT' → Filter 'Foreclosure Cases'"
            }, indent=2)
            
        except Exception as e:
            return json.dumps({
                "status": "error",
                "site": "NYS Court System",
                "error": str(e)
            }, indent=2)


class GreggCountyNavigationTool(BaseTool):
    """Specialized tool for navigating Gregg County Clerk foreclosure data."""
    
    name: str = "gregg_county_navigator"
    description: str = "Navigate Gregg County Clerk site for foreclosure listings"
    args_schema: Type[BaseModel] = WebScraperInput

    def _run(self, url: str = "http://www.co.gregg.tx.us/page/gregg.Clerk.Foreclosure", 
             wait_time: int = 3, css_selectors: Dict[str, str] = None) -> str:
        try:
            headers = {
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
            }
            
            # Access Gregg County Clerk foreclosure page
            response = requests.get(url, headers=headers)
            response.raise_for_status()
            
            soup = BeautifulSoup(response.content, 'html.parser')
            
            # Extract foreclosure listings
            foreclosure_data = []
            
            # Look for foreclosure list or table
            tables = soup.find_all('table')
            for table in tables:
                rows = table.find_all('tr')
                if len(rows) > 1:  # Has header + data
                    # Extract table headers
                    headers_row = rows[0]
                    headers = [th.get_text(strip=True) for th in headers_row.find_all(['th', 'td'])]
                    
                    # Extract data rows
                    for row in rows[1:]:
                        cells = row.find_all(['td', 'th'])
                        if len(cells) == len(headers):
                            row_data = {}
                            for i, cell in enumerate(cells):
                                if i < len(headers):
                                    row_data[headers[i]] = cell.get_text(strip=True)
                            foreclosure_data.append(row_data)
            
            # Look for foreclosure list links
            foreclosure_links = []
            for link in soup.find_all('a', href=True):
                link_text = link.get_text(strip=True).lower()
                if any(keyword in link_text for keyword in ['foreclosure', 'notice', 'list', 'sale']):
                    foreclosure_links.append({
                        "text": link.get_text(strip=True),
                        "href": link['href']
                    })
            
            return json.dumps({
                "status": "success",
                "site": "Gregg County Clerk",
                "county": "Gregg, TX",
                "foreclosure_data": foreclosure_data[:10],  # First 10 records
                "foreclosure_links": foreclosure_links,
                "data_count": len(foreclosure_data),
                "instructions": "If Situs Address missing, check https://esearch.gcad.org/ for property details"
            }, indent=2)
            
        except Exception as e:
            return json.dumps({
                "status": "error",
                "site": "Gregg County Clerk",
                "error": str(e),
                "fallback": "Try accessing foreclosure list directly or check county clerk main page"
            }, indent=2)


class DataExtractorTool(BaseTool):
    """Extract structured data from HTML and map to ATTOM template."""
    
    name: str = "data_extractor"
    description: str = "Extract and structure foreclosure data for ATTOM template"
    args_schema: Type[BaseModel] = DataExtractorInput

    def _run(self, html_content: str, extraction_rules: Dict[str, str]) -> str:
        try:
            soup = BeautifulSoup(html_content, 'html.parser')
            
            extracted_data = {}
            
            for field_name, rule in extraction_rules.items():
                if rule.startswith('css:'):
                    # CSS selector extraction
                    selector = rule[4:]  # Remove 'css:' prefix
                    elements = soup.select(selector)
                    if elements:
                        extracted_data[field_name] = [elem.get_text(strip=True) for elem in elements]
                    else:
                        extracted_data[field_name] = []
                        
                elif rule.startswith('regex:'):
                    # Regex extraction
                    pattern = rule[6:]  # Remove 'regex:' prefix
                    text_content = soup.get_text()
                    matches = re.findall(pattern, text_content, re.IGNORECASE)
                    extracted_data[field_name] = matches
                    
                elif rule.startswith('text:'):
                    # Text-based search
                    search_text = rule[5:]  # Remove 'text:' prefix
                    text_content = soup.get_text()
                    if search_text.lower() in text_content.lower():
                        # Extract surrounding context
                        start_idx = text_content.lower().find(search_text.lower())
                        context = text_content[max(0, start_idx-100):start_idx+200]
                        extracted_data[field_name] = [context.strip()]
                    else:
                        extracted_data[field_name] = []
            
            return json.dumps({
                "status": "success",
                "extracted_fields": extracted_data,
                "field_count": len([k for k, v in extracted_data.items() if v])
            }, indent=2)
            
        except Exception as e:
            return json.dumps({
                "status": "error",
                "error": str(e)
            }, indent=2)


class ExcelWriterInput(BaseModel):
    """Input schema for Excel data writing."""
    data: List[Dict[str, Any]] = Field(..., description="List of records to write to Excel")
    template_path: str = Field(default="docs/ATTOM Foreclosure Template V6.7.xls", description="Path to Excel template file")
    output_path: str = Field(default="output/foreclosure_data.xlsx", description="Path for output Excel file")
    sheet_name: str = Field(default="Sheet1", description="Name of the worksheet to write to")


class HTMLReaderInput(BaseModel):
    """Input schema for reading HTML files."""
    file_path: str = Field(default="search_results_debug.html", description="Path to HTML file to read")


class JSONWriterInput(BaseModel):
    """Input schema for writing JSON files."""
    data: List[Dict[str, Any]] = Field(..., description="List of records to write to JSON file")
    start_date: str = Field(..., description="Start date of search range (MM/DD/YYYY)")
    end_date: str = Field(..., description="End date of search range (MM/DD/YYYY)")
    county: str = Field(default="King", description="County name for the search")
    state: str = Field(default="WA", description="State code")


class AddressVerifierInput(BaseModel):
    """Input schema for address verification from Google Maps."""
    address_link: str = Field(..., description="Google Maps link to verify address")


class JSONReaderInput(BaseModel):
    """Input schema for reading JSON files."""
    file_path: str = Field(..., description="Path to JSON file to read")


class JSONWriterTool(BaseTool):
    """Tool for writing foreclosure data to JSON files with date-based naming."""
    
    name: str = "json_writer"
    description: str = "Save foreclosure records to JSON file with date range in filename"
    args_schema: Type[BaseModel] = JSONWriterInput

    def _run(self, data: List[Dict[str, Any]], start_date: str, end_date: str, 
             county: str = "King", state: str = "WA") -> str:
        try:
            # Get base directory
            base_dir = os.path.dirname(os.path.abspath(__file__))
            
            # Create json_results folder
            json_folder = os.path.join(base_dir, "json_results")
            os.makedirs(json_folder, exist_ok=True)
            
            # Convert dates to filename-friendly format (MM-DD-YYYY)
            start_formatted = start_date.replace('/', '-')
            end_formatted = end_date.replace('/', '-')
            
            # Create filename: foreclosures_WA_King_01-01-2026_to_01-31-2026.json
            filename = f"foreclosures_{state}_{county}_{start_formatted}_to_{end_formatted}.json"
            full_path = os.path.join(json_folder, filename)
            
            # Prepare JSON data with metadata
            json_output = {
                "metadata": {
                    "search_date": datetime.now().isoformat(),
                    "state": state,
                    "county": county,
                    "start_date": start_date,
                    "end_date": end_date,
                    "total_records": len(data),
                    "record_type": "T - Trustee Sale"
                },
                "records": data
            }
            
            # Write JSON file with nice formatting
            with open(full_path, 'w', encoding='utf-8') as f:
                json.dump(json_output, f, indent=2, ensure_ascii=False)
            
            print(f"✅ JSON file saved: {filename}")
            print(f"📊 Total records: {len(data)}")
            print(f"📁 Location: {json_folder}")
            
            return json.dumps({
                "status": "success",
                "message": f"Successfully saved {len(data)} records to JSON file",
                "file_path": full_path,
                "filename": filename,
                "folder": json_folder,
                "records_count": len(data),
                "search_parameters": {
                    "state": state,
                    "county": county,
                    "start_date": start_date,
                    "end_date": end_date
                },
                "timestamp": datetime.now().isoformat()
            }, indent=2)
            
        except Exception as e:
            return json.dumps({
                "status": "error",
                "error": f"Failed to write JSON file: {str(e)}",
                "start_date": start_date,
                "end_date": end_date
            }, indent=2)


class AddressVerifierTool(BaseTool):
    """Tool for verifying addresses by extracting accurate information from Google Maps links."""
    
    name: str = "address_verifier"
    description: str = "Extract accurate address information from Google Maps links"
    args_schema: Type[BaseModel] = AddressVerifierInput

    def _run(self, address_link: str) -> str:
        try:
            headers = {
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
                'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
                'Accept-Language': 'en-US,en;q=0.5',
            }
            
            # Add delay to avoid rate limiting
            time.sleep(random.uniform(2, 4))
            
            # Parse address from URL query parameter (Google Maps links)
            from urllib.parse import urlparse, parse_qs, unquote
            
            verified_address = None
            verified_city = None
            verified_state = None
            verified_zip = None
            
            # Google Maps links format: http://maps.google.com/maps?q=ADDRESS+CITY,+STATE+ZIP
            if 'q=' in address_link:
                parsed = urlparse(address_link)
                query_params = parse_qs(parsed.query)
                if 'q' in query_params:
                    # Decode URL-encoded address
                    address_from_url = unquote(query_params['q'][0])
                    # Replace + with spaces
                    address_from_url = address_from_url.replace('+', ' ')
                    
                    # Parse the address components
                    # Format examples:
                    # "1810 SOUTH 261ST PLACEDES MOINES, WASHINGTON 98198"
                    # "12217 SE PETROVITSKY RDRENTON, WA 98058"
                    
                    # Split by comma first
                    if ',' in address_from_url:
                        parts = address_from_url.split(',')
                        street_city = parts[0].strip()
                        state_zip = parts[1].strip() if len(parts) > 1 else ""
                        
                        # Try to separate street from city in the first part
                        # Look for common city names or patterns
                        # For now, we'll try to extract based on patterns
                        
                        # Extract state and zip from second part
                        if state_zip:
                            state_zip_match = re.search(r'([A-Z]{2,})\s+(\d{5})', state_zip)
                            if state_zip_match:
                                verified_state = state_zip_match.group(1)
                                verified_zip = state_zip_match.group(2)
                        
                        # For street/city separation, look for city names in the first part
                        # Common WA cities
                        wa_cities = ['SEATTLE', 'RENTON', 'DES MOINES', 'AUBURN', 'ENUMCLAW', 
                                    'KIRKLAND', 'KENT', 'BELLEVUE', 'TACOMA', 'FEDERAL WAY']
                        
                        for city in wa_cities:
                            if city in street_city.upper():
                                # Found city in the string
                                city_index = street_city.upper().index(city)
                                verified_address = street_city[:city_index].strip()
                                verified_city = city
                                break
                        
                        # If no city found, use the whole first part as address
                        if not verified_address:
                            verified_address = street_city
                    else:
                        # No comma, use whole thing as address
                        verified_address = address_from_url
            
            # Try fetching the page for additional verification
            try:
                response = requests.get(address_link, headers=headers, timeout=10)
                response.raise_for_status()
                soup = BeautifulSoup(response.content, 'html.parser')
                
                # Try to find address in meta tags
                meta_title = soup.find('meta', {'property': 'og:title'})
                if meta_title and meta_title.get('content'):
                    page_address = meta_title.get('content')
                    # Use this as fallback or verification
                    if not verified_address:
                        verified_address = page_address
            except:
                # If page fetch fails, use URL parsing only
                pass
            
            return json.dumps({
                "status": "success",
                "original_link": address_link,
                "verified_address": verified_address or "Could not extract",
                "verified_city": verified_city or "Could not extract",
                "verified_state": verified_state or "Could not extract",
                "verified_zip": verified_zip or "Could not extract",
                "full_address": f"{verified_address}, {verified_city}, {verified_state} {verified_zip}" if all([verified_address, verified_city, verified_state, verified_zip]) else verified_address or "Extraction incomplete"
            }, indent=2)
            
        except Exception as e:
            return json.dumps({
                "status": "error",
                "error": f"Failed to verify address: {str(e)}",
                "original_link": address_link
            }, indent=2)


class JSONReaderTool(BaseTool):
    """Tool for reading JSON files and returning the data."""
    
    name: str = "json_reader"
    description: str = "Read JSON file with foreclosure records for Excel export"
    args_schema: Type[BaseModel] = JSONReaderInput

    def _run(self, file_path: str) -> str:
        try:
            # Get absolute path
            base_dir = os.path.dirname(os.path.abspath(__file__))
            
            # Handle relative paths
            if not os.path.isabs(file_path):
                full_path = os.path.join(base_dir, file_path)
            else:
                full_path = file_path
            
            # Check if file exists
            if not os.path.exists(full_path):
                return json.dumps({
                    "status": "error",
                    "error": f"JSON file not found. The data_mapper agent may not have completed successfully.",
                    "file_path": file_path,
                    "full_path": full_path,
                    "suggestion": "Ensure the mapping task completed and saved the JSON file before running Excel generation."
                }, indent=2)
            
            # Read JSON file
            with open(full_path, 'r', encoding='utf-8') as f:
                json_data = json.load(f)
            
            # Extract records
            records = json_data.get("records", [])
            metadata = json_data.get("metadata", {})
            
            return json.dumps({
                "status": "success",
                "file_path": full_path,
                "metadata": metadata,
                "total_records": len(records),
                "records": records
            }, indent=2)
            
        except Exception as e:
            return json.dumps({
                "status": "error",
                "error": f"Failed to read JSON file: {str(e)}",
                "file_path": file_path
            }, indent=2)


class HTMLReaderTool(BaseTool):
    """Tool for reading saved HTML files and returning the content."""
    
    name: str = "html_reader"
    description: str = "Read HTML file content from search results for parsing and data extraction"
    args_schema: Type[BaseModel] = HTMLReaderInput

    def _run(self, file_path: str = "search_results_debug.html") -> str:
        try:
            # Get absolute path
            base_dir = os.path.dirname(os.path.abspath(__file__))
            full_path = os.path.join(base_dir, file_path)
            
            # Read HTML file
            with open(full_path, 'r', encoding='utf-8') as f:
                html_content = f.read()
            
            # Parse with BeautifulSoup
            soup = BeautifulSoup(html_content, 'html.parser')
            
            # Find the results table
            tables = soup.find_all('table')
            
            if not tables:
                return json.dumps({
                    "status": "error",
                    "message": "No tables found in HTML file"
                })
            
            # Get the main results table (usually the one with class 'items')
            results_table = None
            for table in tables:
                if 'items' in table.get('class', []):
                    results_table = table
                    break
            
            if not results_table:
                results_table = tables[0]  # Fallback to first table
            
            # Extract table structure
            rows = results_table.find_all('tr')
            
            # Get headers
            header_row = rows[0] if rows else None
            headers = []
            if header_row:
                headers = [th.get_text(strip=True) for th in header_row.find_all(['th', 'td'])]
            
            # Extract all data rows
            data_rows = []
            for row in rows[1:]:  # Skip header row
                cells = row.find_all(['td', 'th'])
                row_data = {}
                
                for i, cell in enumerate(cells):
                    if i < len(headers):
                        # Get cell text (including line breaks preserved)
                        cell_text = cell.get_text(separator='\n', strip=False).strip()
                        row_data[headers[i]] = cell_text
                        
                        # Check for links
                        link = cell.find('a', href=True)
                        if link:
                            row_data[f"{headers[i]}_Link"] = link.get('href')
                
                if row_data:  # Only add non-empty rows
                    data_rows.append(row_data)
            
            return json.dumps({
                "status": "success",
                "file_path": full_path,
                "headers": headers,
                "total_rows": len(data_rows),
                "data": data_rows
            }, indent=2, ensure_ascii=False)
            
        except FileNotFoundError:
            return json.dumps({
                "status": "error",
                "message": f"HTML file not found: {file_path}",
                "full_path": full_path
            })
        except Exception as e:
            return json.dumps({
                "status": "error",
                "message": str(e)
            })


class AttomExcelWriterTool(BaseTool):
    """Tool for writing foreclosure data to ATTOM Excel template with proper field mapping."""
    
    name: str = "excel_writer"
    description: str = "Write foreclosure data to ATTOM Excel template with proper formatting and field mapping according to WA Trustee Sale requirements"
    args_schema: Type[BaseModel] = ExcelWriterInput

    def _run(self, data: List[Dict[str, Any]], template_path: str = "docs/ATTOM Foreclosure Template V6.7.xls", 
             output_path: str = "output/foreclosure_data.xlsx", sheet_name: str = "Sheet1") -> str:
        try:
            # Get absolute paths
            base_dir = os.path.dirname(os.path.abspath(__file__))
            template_full_path = os.path.join(base_dir, template_path)
            output_full_path = os.path.join(base_dir, output_path)
            
            # Create output directory if it doesn't exist
            os.makedirs(os.path.dirname(output_full_path), exist_ok=True)
            
            print(f"📂 Creating Excel file from template: {template_path}")
            
            # Handle .xls format - read headers and create new .xlsx workbook
            if template_full_path.endswith('.xls') and os.path.exists(template_full_path):
                try:
                    import xlrd
                    from openpyxl import Workbook
                    
                    print(f"⚠️ Reading .xls template to extract headers...")
                    
                    # Read .xls file
                    xls_book = xlrd.open_workbook(template_full_path, formatting_info=False)
                    xls_sheet = xls_book.sheet_by_index(0)
                    
                    # Create new .xlsx workbook
                    workbook = Workbook()
                    worksheet = workbook.active
                    
                    # Copy ALL rows from template (including headers and any existing formatting rows)
                    print(f"📋 Copying {xls_sheet.nrows} header rows from template...")
                    for row_idx in range(xls_sheet.nrows):
                        for col_idx in range(xls_sheet.ncols):
                            cell_value = xls_sheet.cell_value(row_idx, col_idx)
                            worksheet.cell(row=row_idx + 1, column=col_idx + 1, value=cell_value)
                    
                    # Apply header formatting to first row
                    if xls_sheet.nrows > 0:
                        for col_idx in range(xls_sheet.ncols):
                            cell = worksheet.cell(row=1, column=col_idx + 1)
                            cell.font = Font(bold=True)
                            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                            cell.alignment = Alignment(horizontal="center")
                    
                    # Store the number of template rows
                    template_rows = xls_sheet.nrows
                    print(f"✅ Template loaded with {template_rows} rows, {xls_sheet.ncols} columns")
                    
                except ImportError:
                    return json.dumps({
                        "status": "error",
                        "error": "xlrd library not installed. Please install it: pip install xlrd",
                        "template_path": template_full_path
                    }, indent=2)
                    
            elif os.path.exists(template_full_path):
                workbook = openpyxl.load_workbook(template_full_path)
                worksheet = workbook.active
                template_rows = worksheet.max_row
                print(f"✅ Loaded .xlsx template: {template_path}")
            else:
                return json.dumps({
                    "status": "error",
                    "error": f"Template file not found: {template_full_path}"
                }, indent=2)
            
            # Get starting row (after template headers/formatting rows)
            start_row = template_rows + 1
            print(f"📝 Writing {len(data)} records starting at row {start_row}")
            
            # Field mapping based on ATTOM Foreclosure Template V6.7
            # Column positions based on actual template structure:
            # A: RECORD TYPE, B: RECORDING DATE, C: JUDGMENT DATE, D: CASE NO,
            # E: FORECLOSURE DOC NO, F: LOAN NO, G: TRUSTEE SALE NO, H: DATE OF LOAN,
            # I: DEED OF TRUST DOC NO, J: MORTGAGE VOLUME PAGE, K: LOAN AMOUNT, L: INT RATE,
            # M: LOAN MATURITY DATE, N: PAYMENT, O: BORROWER OR DEFENDANT NAME,
            # P: LENDER OR PLAINTIFF NAME, Q-T: LENDER ADDRESS/CITY/STATE/ZIP, U: LENDER PHONE NO,
            # V-Y: SERVICER NAME/ADDRESS/CITY/STATE, Z: SERVICER ZIP, AA: SERVICER PHONE NO,
            # AB: TRUSTEE OR DEPUTY NAME, AC-AF: TRUSTEE ADDRESS/CITY/STATE/ZIP, AG: TRUSTEE PHONE NO,
            # AH: BUYER NAME, AI: PROPERTY DESCRIPTION, AJ-AN: SITUS ADDRESS/CITY/COUNTY/STATE/ZIP,
            # AO: PARCEL ID NO, AP: APPRAISED VALUE, AQ: MINIMUM BID, AR: DEFAULT AMOUNT,
            # AS: JUDGMENT AMOUNT, AT: LOAN BALANCE, AU: PURCHASE AMOUNT, AV: PENALTIES AND INTEREST,
            # AW: AUCTION DATE, AX: AUCTION TIME, AY-AZ: AUCTION ADDRESS/CITY
            
            for row_idx, record in enumerate(data, start_row):
                # Column A (1): RECORD TYPE - "T" for WA Trustee Sales
                worksheet.cell(row=row_idx, column=1, value=record.get("Record_Type", "T"))
                
                # Column B (2): RECORDING DATE
                worksheet.cell(row=row_idx, column=2, value=record.get("Recording_Date", ""))
                
                # Column C (3): JUDGMENT DATE
                worksheet.cell(row=row_idx, column=3, value="")
                
                # Column D (4): CASE NO
                worksheet.cell(row=row_idx, column=4, value="")
                
                # Column E (5): FORECLOSURE DOC NO - EPP Number
                worksheet.cell(row=row_idx, column=5, value=record.get("EPP_Number", ""))
                
                # Column F (6): LOAN NO
                worksheet.cell(row=row_idx, column=6, value="")
                
                # Column G (7): TRUSTEE SALE NO - TS Number
                worksheet.cell(row=row_idx, column=7, value=record.get("TS_Number", ""))
                
                # Columns H-N (8-14): Skip loan details (not available)
                for col in range(8, 15):
                    worksheet.cell(row=row_idx, column=col, value="")
                
                # Column O (15): BORROWER OR DEFENDANT NAME
                worksheet.cell(row=row_idx, column=15, value="")
                
                # Columns P-U (16-21): Skip lender details (not available)
                for col in range(16, 22):
                    worksheet.cell(row=row_idx, column=col, value="")
                
                # Columns V-AA (22-27): Skip servicer details (not available)
                for col in range(22, 28):
                    worksheet.cell(row=row_idx, column=col, value="")
                
                # Column AB (28): TRUSTEE OR DEPUTY NAME
                worksheet.cell(row=row_idx, column=28, value=record.get("Trustee_Name", ""))
                
                # Column AC (29): TRUSTEE ADDRESS (street only)
                worksheet.cell(row=row_idx, column=29, value=record.get("Trustee_Address", ""))
                
                # Column AD (30): TRUSTEE CITY
                worksheet.cell(row=row_idx, column=30, value=record.get("Trustee_City", ""))
                
                # Column AE (31): TRUSTEE STATE
                worksheet.cell(row=row_idx, column=31, value=record.get("Trustee_State", ""))
                
                # Column AF (32): TRUSTEE ZIP
                worksheet.cell(row=row_idx, column=32, value=record.get("Trustee_Zip", ""))
                
                # Column AG (33): TRUSTEE PHONE NO
                worksheet.cell(row=row_idx, column=33, value="")
                
                # Column AH (34): BUYER NAME
                worksheet.cell(row=row_idx, column=34, value="")
                
                # Column AI (35): PROPERTY DESCRIPTION
                worksheet.cell(row=row_idx, column=35, value="")
                
                # Column AJ (36): SITUS ADDRESS (from Google Maps verified address)
                worksheet.cell(row=row_idx, column=36, value=record.get("Situs_Address", ""))
                
                # Column AK (37): SITUS CITY
                worksheet.cell(row=row_idx, column=37, value=record.get("Situs_City", ""))
                
                # Column AL (38): SITUS COUNTY
                worksheet.cell(row=row_idx, column=38, value=record.get("County", "King"))
                
                # Column AM (39): SITUS STATE
                worksheet.cell(row=row_idx, column=39, value=record.get("Situs_State", "WA"))
                
                # Column AN (40): SITUS ZIP
                worksheet.cell(row=row_idx, column=40, value=record.get("Situs_Zip", ""))
                
                # Column AO (41): PARCEL ID NO
                worksheet.cell(row=row_idx, column=41, value="")
                
                # Column AP (42): APPRAISED VALUE - Estimated Sale Amount
                sale_amount = record.get("Estimated_Sale_Amount", "")
                if isinstance(sale_amount, (int, float)):
                    worksheet.cell(row=row_idx, column=42, value=sale_amount)
                else:
                    worksheet.cell(row=row_idx, column=42, value=sale_amount)
                
                # Column AQ (43): MINIMUM BID - Bid Amount
                bid_amount = record.get("Bid_Amount", "")
                if isinstance(bid_amount, (int, float)):
                    worksheet.cell(row=row_idx, column=43, value=bid_amount)
                else:
                    # Clean up bid amount string
                    if bid_amount and bid_amount != "":
                        # Remove commas and convert to number if possible
                        try:
                            clean_bid = str(bid_amount).replace(',', '').replace('$', '').strip()
                            if clean_bid:
                                worksheet.cell(row=row_idx, column=43, value=float(clean_bid))
                            else:
                                worksheet.cell(row=row_idx, column=43, value="")
                        except:
                            worksheet.cell(row=row_idx, column=43, value=bid_amount)
                    else:
                        worksheet.cell(row=row_idx, column=43, value="")
                
                # Column AR (44): DEFAULT AMOUNT
                worksheet.cell(row=row_idx, column=44, value="")
                
                # Column AS (45): JUDGMENT AMOUNT
                worksheet.cell(row=row_idx, column=45, value="")
                
                # Column AT (46): LOAN BALANCE
                worksheet.cell(row=row_idx, column=46, value="")
                
                # Column AU (47): PURCHASE AMOUNT
                worksheet.cell(row=row_idx, column=47, value="")
                
                # Column AV (48): PENALTIES AND INTEREST
                worksheet.cell(row=row_idx, column=48, value="")
                
                # Column AW (49): AUCTION DATE - Sale Date
                worksheet.cell(row=row_idx, column=49, value=record.get("Sale_Date", ""))
                
                # Column AX (50): AUCTION TIME - Sale Time
                worksheet.cell(row=row_idx, column=50, value=record.get("Sale_Time", ""))
                
                # Column AY (51): AUCTION ADDRESS
                worksheet.cell(row=row_idx, column=51, value="")
                
                # Column AZ (52): AUCTION CITY
                worksheet.cell(row=row_idx, column=52, value="")
                
                # Columns BA-BK (53-63): Skip remaining fields (dates, document numbers not available)
                for col in range(53, 64):
                    worksheet.cell(row=row_idx, column=col, value="")
                
                # Add custom tracking columns at the end
                # Column BL (64): Address Link (Google Maps URL)
                worksheet.cell(row=row_idx, column=64, value=record.get("Address_Link", ""))
                
                # Column BM (65): Address Verified Status
                worksheet.cell(row=row_idx, column=65, value=record.get("Address_Verified", "Pending"))
                
                # Column BN (66): CAPTCHA Status
                worksheet.cell(row=row_idx, column=66, value=record.get("CAPTCHA_Status", "Solved"))
                
                # Column BO (67): Status
                worksheet.cell(row=row_idx, column=67, value=record.get("Status", ""))
                
                # Column BP (68): Sold
                worksheet.cell(row=row_idx, column=68, value=record.get("Sold", "No"))
                
                # Column BQ (69): Bid Info
                worksheet.cell(row=row_idx, column=69, value=record.get("Bid_Info", ""))
                
            print(f"✅ Wrote {len(data)} records to Excel")
            
            # Auto-adjust column widths for readability
            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Save workbook
            workbook.save(output_full_path)
            print(f"💾 Saved to: {output_full_path}")
            
            return json.dumps({
                "status": "success",
                "message": f"Successfully wrote {len(data)} records to ATTOM template",
                "records_written": len(data),
                "output_file": output_full_path,
                "template_used": template_full_path,
                "starting_row": start_row,
                "timestamp": datetime.now().isoformat()
            }, indent=2)
            
        except Exception as e:
            import traceback
            return json.dumps({
                "status": "error",
                "error": f"Failed to write Excel file: {str(e)}",
                "traceback": traceback.format_exc(),
                "template_path": template_path,
                "output_path": output_path
            }, indent=2)


class ElitePostingScraperInput(BaseModel):
    """Input schema for Elite Posting & Publishing website scraping."""
    url: str = Field(default="https://elitepostandpub.com/INDEX.PHP", description="Elite Posting & Publishing website URL")
    state: str = Field(default="WA", description="State code (WA for Washington)")
    county: str = Field(default="King", description="County name for foreclosure search")
    start_date: str = Field(default="05/01/2025", description="Start date for sale date range (MM/DD/YYYY)")
    end_date: str = Field(default="05/14/2025", description="End date for sale date range (MM/DD/YYYY)")
    output_file: str = Field(default="search_results_debug.html", description="Output filename for saved HTML results")
    captcha_code: Optional[str] = Field(default=None, description="CAPTCHA verification code if known")

class ElitePostingScraperTool(BaseTool):
    """Real Elite Posting & Publishing website scraper for Washington State foreclosure data."""
    
    name: str = "elite_posting_scraper"
    description: str = "Extract real foreclosure data from Elite Posting & Publishing website with CAPTCHA handling"
    args_schema: Type[BaseModel] = ElitePostingScraperInput

    def _solve_captcha_automatically(self, captcha_file_path):
        """Enhanced CAPTCHA solver using Vision LLM and multiple fallback methods."""
        print(f"🔍 Attempting to solve CAPTCHA from: {captcha_file_path}")
        
        # Use the enhanced solver with Vision LLM + API + Local OCR fallback
        try:
            from captcha_solver import solve_captcha
            
            print("🚀 Using Enhanced CAPTCHA Solver (Vision LLM + APIs + Local OCR)")
            result = solve_captcha(captcha_file_path)
            
            if result and result != "FAILED":
                print(f"✅ CAPTCHA Solved: {result}")
                return result
            else:
                print("❌ Enhanced solver failed")
                
        except Exception as e:
            print(f"❌ Enhanced solver error: {str(e)[:100]}")

        print("❌ CAPTCHA solving failed")
        return "FAILED"
    
    def _apply_captcha_corrections(self, text):
        """Apply common CAPTCHA OCR corrections for random mixed-case CAPTCHAs."""
        import re
        
        corrected = text
        original = text
        corrections_applied = []
        
        # Pattern-based corrections that work for any random CAPTCHA
        
        # 1. Fix "ml" pattern back to "mon" (case-insensitive)
        # Handles: kemlew->kemonew, KEMLEw->KEMONEw, etc.
        ml_pattern = re.compile(r'([a-zA-Z])ml([a-zA-Z])', re.IGNORECASE)
        if ml_pattern.search(corrected):
            def ml_replacement(match):
                prefix = match.group(1)
                suffix = match.group(2)
                # Preserve case of the 'o' based on surrounding letters
                if prefix.islower() and suffix.islower():
                    return f"{prefix}mon{suffix}"
                elif prefix.isupper() and suffix.isupper():
                    return f"{prefix}MON{suffix}"
                else:
                    # Mixed case - preserve original case pattern
                    return f"{prefix}mon{suffix}"
            
            new_corrected = ml_pattern.sub(ml_replacement, corrected)
            if new_corrected != corrected:
                corrections_applied.append(f"ml->mon pattern: {corrected} -> {new_corrected}")
                corrected = new_corrected
        
        # 2. Fix "rn" -> "m" pattern (case-insensitive)
        rn_pattern = re.compile(r'rn', re.IGNORECASE)
        if rn_pattern.search(corrected):
            def rn_replacement(match):
                return 'm' if match.group().islower() else 'M'
            
            new_corrected = rn_pattern.sub(rn_replacement, corrected)
            if new_corrected != corrected:
                corrections_applied.append(f"rn->m pattern: {corrected} -> {new_corrected}")
                corrected = new_corrected
        
        # 3. Fix "cl" -> "d" pattern (case-insensitive)  
        cl_pattern = re.compile(r'cl', re.IGNORECASE)
        if cl_pattern.search(corrected):
            def cl_replacement(match):
                return 'd' if match.group().islower() else 'D'
            
            new_corrected = cl_pattern.sub(cl_replacement, corrected)
            if new_corrected != corrected:
                corrections_applied.append(f"cl->d pattern: {corrected} -> {new_corrected}")
                corrected = new_corrected
        
        # 4. Handle common single character confusions (preserve case)
        char_mapping = {
            'B': 'g', 'b': 'g',  # B/b often confused with g
            '0': 'O', '0': 'o',  # Zero vs O/o  
            '1': 'I', '1': 'i',  # One vs I/i
            '1': 'l', '1': 'L',  # One vs l/L
            '8': 'B', '8': 'b',  # Eight vs B/b
            '6': 'G', '6': 'g',  # Six vs G/g
            '5': 'S', '5': 's',  # Five vs S/s
            '2': 'Z', '2': 'z',  # Two vs Z/z
        }
        
        # Apply character corrections only if it makes sense contextually
        for wrong, right in char_mapping.items():
            if wrong in corrected:
                # Be conservative - only replace if it improves readability
                test_correction = corrected.replace(wrong, right, 1)  # Replace only first occurrence
                corrections_applied.append(f"Character option: {wrong}->{right} ({corrected} -> {test_correction})")
        
        # 5. Handle specific OCR patterns that commonly occur
        # Fix "ij" -> "j" (case-insensitive)
        ij_pattern = re.compile(r'ij', re.IGNORECASE)
        if ij_pattern.search(corrected):
            def ij_replacement(match):
                return 'j' if match.group().islower() else 'J'
            
            new_corrected = ij_pattern.sub(ij_replacement, corrected)
            if new_corrected != corrected:
                corrections_applied.append(f"ij->j pattern: {corrected} -> {new_corrected}")
                corrected = new_corrected
        
        # Log all applied corrections
        if corrections_applied:
            for correction in corrections_applied:
                print(f"Applied: {correction}")
        
        # Only return corrected version if it's meaningfully different and likely better
        if corrected != original and len(corrected) >= len(original) - 1:
            print(f"Final correction: {original} -> {corrected}")
            return corrected
        
        # Return original if no meaningful corrections applied
        return original
    
    def _run(self, url: str = "https://elitepostandpub.com/INDEX.PHP", state: str = "WA", county: str = "King",
             start_date: str = "05/01/2025", end_date: str = "05/14/2025", 
             output_file: str = "search_results_debug.html", captcha_code: str = None) -> str:
        try:
            # Force the correct URL regardless of input
            url = "https://elitepostandpub.com/INDEX.PHP"
            
            # Enhanced headers to avoid detection
            headers = {
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
                'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
                'Accept-Language': 'en-US,en;q=0.5',
                'Accept-Encoding': 'gzip, deflate, br',
                'Connection': 'keep-alive',
                'Upgrade-Insecure-Requests': '1',
                'Sec-Fetch-Dest': 'document',
                'Sec-Fetch-Mode': 'navigate',
                'Sec-Fetch-Site': 'none',
                'Cache-Control': 'max-age=0'
            }
            
            # Create session for cookie persistence
            session = requests.Session()
            session.headers.update(headers)
            
            # Step 1: Access main site
            print(f"Accessing Elite Posting & Publishing website: {url}")
            response = session.get(url)
            response.raise_for_status()
            
            soup = BeautifulSoup(response.content, 'html.parser')
            
            # Step 2: Find search form or navigation to search page
            search_form = soup.find('form') or soup.find('div', {'class': lambda x: x and 'search' in x.lower()})
            
            # Look for search page link if form not found on main page
            search_links = soup.find_all('a', href=True)
            search_url = None
            
            for link in search_links:
                link_text = link.get_text(strip=True).lower()
                href = link['href']
                if any(keyword in link_text for keyword in ['search', 'foreclosure', 'trustee']) or 'search' in href.lower():
                    print(f"Found search link: {link_text} -> {href}")
                    # Fix URL construction - use base domain instead of full URL with INDEX.PHP
                    if href.startswith('http'):
                        search_url = href
                    elif href.startswith('/'):
                        # For absolute paths, use only the domain part
                        base_domain = url.split('/INDEX.PHP')[0]  # Get https://elitepostandpub.com
                        search_url = base_domain + href
                    else:
                        # For relative paths, append to base domain
                        base_domain = url.split('/INDEX.PHP')[0]
                        search_url = base_domain + '/' + href.lstrip('/')
                    break
            
            # Step 3: Navigate to search page
            if search_url:
                print(f"Navigating to search page: {search_url}")
                response = session.get(search_url)
                response.raise_for_status()
                soup = BeautifulSoup(response.content, 'html.parser')
            
            # Step 4: Find the SEARCH form (not the login form)
            # Look for the form with id="orders-search-form" or action containing "orders/search"
            search_form = None
            all_forms = soup.find_all('form')
            
            for form in all_forms:
                form_id = form.get('id', '')
                form_action = form.get('action', '')
                # Find the orders search form specifically
                if 'orders' in form_id.lower() or 'orders' in form_action.lower() or 'search' in form_id.lower():
                    search_form = form
                    print(f"✅ Found search form: id={form_id}, action={form_action}")
                    break
            
            if not search_form and all_forms:
                # If no specific search form found, use the last form (usually the main form on the page)
                search_form = all_forms[-1] if len(all_forms) > 1 else all_forms[0]
                print(f"⚠️ Using fallback form: {search_form.get('id', 'no-id')}, action={search_form.get('action', 'no-action')}")
            
            if not search_form:
                return json.dumps({
                    "status": "error",
                    "message": "Search form not found on Elite Posting website",
                    "available_forms": len(soup.find_all('form')),
                    "page_title": soup.find('title').get_text() if soup.find('title') else "No title"
                })
            
            # Prepare form data
            form_data = {}
            captcha_field_name = None
            
            # Map to actual form field names based on URL analysis
            # https://elitepostandpub.com/INDEX.PHP?r=orders%2Fsearch
            # Orders[PropertyStateCode]=WA
            # Orders[PropertyCountyID]=94 (King County)
            # Orders[startSaleDateTime]=01/01/2026
            # Orders[endSaleDateTime]=01/31/2026
            # Orders[verifyCode]=doyoka
            # yt0=Search
            
            # Set the exact form fields
            form_data['Orders[PropertyStateCode]'] = state
            form_data['Orders[PropertyCountyID]'] = '94'  # King County ID (hardcoded)
            form_data['Orders[startSaleDateTime]'] = start_date
            form_data['Orders[endSaleDateTime]'] = end_date
            form_data['Orders[id]'] = ''  # Empty EPP#
            form_data['Orders[TsNumber]'] = ''  # Empty TS#
            form_data['Orders[APN]'] = ''  # Empty APN
            form_data['Orders[PropertyZip]'] = ''  # Empty Zip
            form_data['yt0'] = 'Search'  # Submit button
            
            print(f"📍 County hardcoded to King (ID: 94)")
            
            # Find CAPTCHA field - should be Orders[verifyCode]
            captcha_field_name = 'Orders[verifyCode]'
            
            # Also check for any hidden fields in the form
            for input_field in search_form.find_all(['input', 'select']):
                field_name = input_field.get('name', '')
                field_type = input_field.get('type', 'text')
                
                if field_type == 'hidden':
                    form_data[field_name] = input_field.get('value', '')
            
            # Step 5: Handle CAPTCHA if present
            captcha_img = soup.find('img', {'alt': lambda x: x and 'captcha' in x.lower()}) or \
                         soup.find('img', {'src': lambda x: x and 'captcha' in x.lower()})
            
            if captcha_img and not captcha_code:
                captcha_src = captcha_img.get('src', '')
                
                # Fix CAPTCHA image URL construction
                if captcha_src.startswith('/'):
                    base_domain = url.split('/INDEX.PHP')[0]
                    captcha_full_url = base_domain + captcha_src
                else:
                    captcha_full_url = captcha_src
                
                # Try to download and solve CAPTCHA automatically
                captcha_saved = False
                captcha_file_path = None
                auto_solved_code = None
                
                try:
                    captcha_response = session.get(captcha_full_url)
                    if captcha_response.status_code == 200:
                        # Save CAPTCHA image to file
                        captcha_file_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'captcha_image.png')
                        with open(captcha_file_path, 'wb') as f:
                            f.write(captcha_response.content)
                        captcha_saved = True
                        print(f"CAPTCHA image saved to: {captcha_file_path}")
                        
                        # Simple direct CAPTCHA solving
                        auto_solved_code = self._solve_captcha_automatically(captcha_file_path)
                        
                        if auto_solved_code and auto_solved_code != "FAILED" and len(auto_solved_code) >= 3:
                            # Apply corrections
                            corrected_code = self._apply_captcha_corrections(auto_solved_code)
                            if corrected_code != auto_solved_code:
                                print(f"🔧 Applied correction: {auto_solved_code} → {corrected_code}")
                                auto_solved_code = corrected_code
                            
                            # Set CAPTCHA in the correct field: Orders[verifyCode]
                            form_data['Orders[verifyCode]'] = auto_solved_code
                            captcha_code = auto_solved_code
                            print(f"✅ CAPTCHA set in Orders[verifyCode]: {auto_solved_code}")
                        else:
                            print(f"❌ CAPTCHA solving failed: {auto_solved_code}")
                            
                except Exception as e:
                    print(f"Could not download CAPTCHA image: {e}")
                    return json.dumps({
                        "status": "debug_error",
                        "message": "Failed to download CAPTCHA image",
                        "error": str(e)
                    })
                
                # This will never be reached because we return early above
                # Continue with form submission - we have a CAPTCHA code (either solved or fallback)
                if captcha_code:
                    print(f"✅ Proceeding with CAPTCHA code: {captcha_code}")
                else:
                    print("⚠️ No CAPTCHA code available, proceeding without CAPTCHA")
            
            # Step 6: Review all form fields before submission
            print("\n" + "="*70)
            print("📋 FORM DATA REVIEW BEFORE SUBMISSION")
            print("="*70)
            print(f"State (Orders[PropertyStateCode]): {form_data.get('Orders[PropertyStateCode]', 'N/A')}")
            print(f"County ID (Orders[PropertyCountyID]): {form_data.get('Orders[PropertyCountyID]', 'N/A')} (King=94)")
            print(f"Start Date (Orders[startSaleDateTime]): {form_data.get('Orders[startSaleDateTime]', 'N/A')}")
            print(f"End Date (Orders[endSaleDateTime]): {form_data.get('Orders[endSaleDateTime]', 'N/A')}")
            print(f"CAPTCHA (Orders[verifyCode]): {form_data.get('Orders[verifyCode]', 'Not set')}")
            print(f"\n✅ All form fields being submitted:")
            for field_name, field_value in form_data.items():
                print(f"  {field_name} = {field_value}")
            print("="*70 + "\n")
            
            # Step 7: Submit search form
            # ALWAYS use the correct search URL regardless of form action
            base_domain = url.split('/INDEX.PHP')[0]
            form_action = base_domain + '/INDEX.PHP?r=orders/search'
            
            print(f"🎯 Using correct search URL: {form_action}")
            print(f"Method: GET (parameters in URL)")
            
            # Prepare request to show complete URL
            from requests import Request
            prepared_request = Request('GET', form_action, params=form_data).prepare()
            print(f"\n📍 COMPLETE URL BEING SENT:")
            print(f"{prepared_request.url}")
            print("="*100 + "\n")
            
            print("🚀 Clicking SEARCH button...\n")
            
            # Use GET request with parameters in URL (like the browser does)
            response = session.get(form_action, params=form_data)
            
            response.raise_for_status()
            print(f"✅ Search submitted successfully! Status code: {response.status_code}")
            print(f"Response URL: {response.url}")
            print(f"Response history: {response.history}")
            
            # Parse first page
            results_soup = BeautifulSoup(response.content, 'html.parser')
            
            # Check for pagination
            summary = results_soup.find('div', class_='summary')
            total_records = 0
            current_page = 1
            records_per_page = 10
            
            if summary:
                summary_text = summary.get_text()
                print(f"📊 {summary_text}")
                
                # Extract total count from "Displaying 1-10 of 14 result(s)."
                import re
                match = re.search(r'of\s+(\d+)\s+result', summary_text)
                if match:
                    total_records = int(match.group(1))
                    total_pages = (total_records + records_per_page - 1) // records_per_page
                    print(f"📄 Total records: {total_records}, Total pages: {total_pages}")
            
            # Collect all page HTMLs
            all_pages_html = [response.text]
            
            # If there are multiple pages, fetch them all
            if total_records > records_per_page:
                total_pages = (total_records + records_per_page - 1) // records_per_page
                print(f"\n🔄 Fetching all {total_pages} pages...")
                
                for page_num in range(2, total_pages + 1):
                    print(f"   📄 Fetching page {page_num}...")
                    
                    # Add page parameter to form data
                    page_form_data = form_data.copy()
                    page_form_data['Orders_page'] = str(page_num)
                    
                    try:
                        page_response = session.get(form_action, params=page_form_data)
                        page_response.raise_for_status()
                        all_pages_html.append(page_response.text)
                        print(f"   ✅ Page {page_num} fetched successfully")
                        
                        # Small delay to avoid rate limiting
                        time.sleep(1)
                    except Exception as e:
                        print(f"   ❌ Failed to fetch page {page_num}: {str(e)[:100]}")
            
            # Combine all pages into single HTML for easier processing
            # We'll merge all table rows into the first page's HTML
            if len(all_pages_html) > 1:
                print(f"\n🔗 Merging {len(all_pages_html)} pages into single HTML...")
                
                # Parse first page
                combined_soup = BeautifulSoup(all_pages_html[0], 'html.parser')
                main_table = combined_soup.find('table', class_='items')
                
                if main_table:
                    # Get tbody or create one
                    tbody = main_table.find('tbody')
                    if not tbody:
                        tbody = combined_soup.new_tag('tbody')
                        # Move existing rows to tbody
                        for row in main_table.find_all('tr')[1:]:  # Skip header
                            tbody.append(row.extract())
                        main_table.append(tbody)
                    
                    # Add rows from other pages
                    for page_html in all_pages_html[1:]:
                        page_soup = BeautifulSoup(page_html, 'html.parser')
                        page_table = page_soup.find('table', class_='items')
                        
                        if page_table:
                            # Extract data rows (skip header)
                            page_rows = page_table.find_all('tr')[1:]
                            for row in page_rows:
                                tbody.append(row)
                    
                    # Update summary to reflect all records
                    if summary:
                        summary.string = f"Displaying 1-{total_records} of {total_records} result(s). (All pages merged)"
                    
                    combined_html = str(combined_soup)
                    print(f"✅ Successfully merged all pages")
                else:
                    combined_html = all_pages_html[0]
                    print(f"⚠️ Could not find table to merge, using first page only")
            else:
                combined_html = all_pages_html[0]
            
            # Save combined HTML with unique filename for this date range
            debug_html_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), output_file)
            with open(debug_html_path, 'w', encoding='utf-8') as f:
                f.write(combined_html)
            print(f"📄 HTML saved to: {debug_html_path}")
            print(f"📊 HTML length: {len(combined_html)} characters\n")
            
            results_soup = BeautifulSoup(combined_html, 'html.parser')
            
            # Debug: Check what we received
            page_title = results_soup.find('title')
            print(f"📄 Page Title: {page_title.get_text() if page_title else 'No title'}")
            
            # Check for error messages
            error_msgs = results_soup.find_all(['div', 'span'], class_=lambda x: x and ('error' in str(x).lower() or 'alert' in str(x).lower()))
            if error_msgs:
                print(f"⚠️ Found {len(error_msgs)} error/alert messages:")
                for msg in error_msgs[:3]:
                    print(f"   - {msg.get_text(strip=True)[:200]}")
            
            # Step 7: Quick analysis of response structure (for agent's reference)
            # DO NOT parse or map data - that's the data_mapper agent's job
            
            tables = results_soup.find_all('table')
            print(f"🔍 Found {len(tables)} table(s) in response")
            
            table_info = []
            for idx, table in enumerate(tables):
                rows = table.find_all('tr')
                print(f"   Table {idx+1}: {len(rows)} rows")
                
                if len(rows) > 1:
                    # Get headers for reference
                    header_row = rows[0]
                    headers = [th.get_text(strip=True) for th in header_row.find_all(['th', 'td'])]
                    print(f"   Headers: {headers}")
                    
                    table_info.append({
                        "table_index": idx + 1,
                        "total_rows": len(rows),
                        "data_rows": len(rows) - 1,  # Excluding header
                        "headers": headers
                    })
            
            # Return summary for data_mapper agent to process
            return json.dumps({
                "status": "success",
                "site": "Elite Posting & Publishing",
                "search_parameters": {
                    "state": state,
                    "county": county,
                    "start_date": start_date,
                    "end_date": end_date
                },
                "html_file_saved": debug_html_path,
                "html_file_size_bytes": len(response.text),
                "page_title": page_title.get_text() if page_title else "No title",
                "tables_found": len(tables),
                "table_structure": table_info,
                "captcha_status": "Solved" if captcha_code else "Not required",
                "message": "HTML response saved successfully. Data extraction and field mapping will be handled by data_mapper agent."
            }, indent=2)
            
        except Exception as e:
            return json.dumps({
                "status": "error",
                "site": "Elite Posting & Publishing",
                "error": str(e),
                "error_type": type(e).__name__,
                "troubleshooting": {
                    "check_url": "Verify Elite Posting & Publishing website is accessible",
                    "check_structure": "Website structure may have changed",
                    "check_captcha": "CAPTCHA may be required for access"
                }
            }, indent=2)


# Tool instances to be imported
web_scraper = BasicWebScrapingTool()
nys_court_navigator = NYSCourtNavigationTool()
gregg_county_navigator = GreggCountyNavigationTool()
data_extractor = DataExtractorTool()
excel_writer = AttomExcelWriterTool()
elite_posting_scraper = ElitePostingScraperTool()
html_reader = HTMLReaderTool()
json_writer = JSONWriterTool()
address_verifier = AddressVerifierTool()
json_reader = JSONReaderTool()